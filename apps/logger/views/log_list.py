from datetime import datetime

from django.db.models import Count, Sum
from django.http import HttpResponse
from django_filters.views import FilterView
from openpyxl import Workbook

from ..filters import LogListFilter
from ..models import Log


class LogListView(FilterView):
    """Log List View"""

    model = Log
    paginate_by = 10
    filterset_class = LogListFilter
    template_name = 'logger/log_list.html'

    def get_initial(self):
        """Define default initial"""
        return {
            'search': self.search
        }

    def get_queryset(self):
        """Override queryset"""
        return LogListFilter(self.request.GET, queryset=self.model.objects.all()).qs

    def get_context_data(self, **kwargs):
        """Override template context"""
        context = super().get_context_data(**kwargs)
        paginator = context['paginator']
        page_obj = context['page_obj']

        index = page_obj.number - 1
        max_index = paginator.count

        start_index = index - 3 if index >= 3 else 0
        end_index = index + 3 if index <= max_index - 3 else max_index
        page_range = paginator.page_range[start_index:end_index]

        context['page_range'] = page_range

        queryset = self.get_queryset()

        # aggreagation
        context['unique_ip_address_count'] = queryset.order_by('ip_address').distinct('ip_address').count()
        context['ip_address_counts'] = queryset.values('ip_address').annotate(total=Count('ip_address')) \
            .order_by('-total')[:10]
        context['method_counts'] = queryset.values('method').annotate(total=Count('method')).order_by('-total')
        context['total_size'] = queryset.aggregate(total_size=Sum('size')).get('total_size')

        return context

    def export_xlsx(self, request):
        """Export xlsx"""
        queryset = self.get_queryset()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-logs.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Logs'

        # Define the titles for columns
        columns = [
            'method',
            'code',
            'ip_address',
            'uri',
            'created_at',
            'size',
        ]

        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all logs
        for log in queryset:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                log.method,
                log.code,
                log.ip_address,
                log.uri,
                log.created_at,
                log.size,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response

    def get(self, request, *args, **kwargs):
        """Override get method with xlsx export parameter"""
        if request.GET.get('export_xlsx') == 'yes':
            return self.export_xlsx(request)
        return super().get(request, *args, **kwargs)
